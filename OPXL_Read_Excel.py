from openpyxl import load_workbook

wb = load_workbook("sample.xlsx")
ws = wb.active

# For specific value
print("A1:", ws["A1"].value)
print("B1:", ws["B1"].value)

for row in ws.iter_rows(values_only=True):
    print(row)
