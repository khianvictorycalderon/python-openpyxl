from openpyxl import Workbook

wb = Workbook()
ws = wb.active

ws.title = "MySheet"

ws["A1"] = "Hello"
ws["B1"] = "World"
ws["A2"] = "I"
ws["A3"] = "am"
ws["A4"] = "Khian"
ws.append([1, 2, 3]) # Appends data to the next empty row after the last non-empty row

wb.save("sample.xlsx")
print("Excel file 'Sample.xlsx' created successfully.")